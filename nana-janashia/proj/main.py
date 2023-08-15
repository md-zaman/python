import openpyxl

inv_file = openpyxl.load_workbook("inventory.xls")
product_list = inv_file["Sheet1"]

products_per_supplier = {}

print(product_list.max_row)

# for product_row in product_list.max_row